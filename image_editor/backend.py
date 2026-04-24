import base64
import io
import json
import os
import re
import time
import urllib.parse
import urllib.request

import streamlit as st

DRIVE_FOLDER_ID = os.environ.get('DRIVE_FOLDER_ID', '1BXKQxbRIR1m9rsVNCbmpFk1dDTUke4WO')
SCOPES = ['https://www.googleapis.com/auth/drive']


class DriveUploadError(Exception):
    pass


def load_xlsx(file_bytes: bytes):
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    products = []
    row_map = {}
    for row_idx, vals in enumerate(ws.iter_rows(min_row=2, max_col=91, values_only=True), start=2):
        prod_no = vals[0] if len(vals) > 0 else None
        if not prod_no:
            continue

        prod_no = str(prod_no).strip()
        row_map[prod_no] = row_idx
        prod_name = str(vals[7] if len(vals) > 7 and vals[7] is not None else '').strip()
        img_cl = str(vals[89] if len(vals) > 89 and vals[89] is not None else '').strip()
        img_cm = str(vals[90] if len(vals) > 90 and vals[90] is not None else '').strip()

        detail_html = str(vals[53] if len(vals) > 53 and vals[53] is not None else '')
        detail_imgs = _extract_img_urls(detail_html)

        products.append({
            'prod_no': prod_no,
            'prod_name': prod_name,
            'img_cl': img_cl,
            'img_cm': img_cm,
            'detail_imgs': detail_imgs,
        })

    wb.close()
    return products, row_map


def _extract_img_urls(html_str: str):
    if not html_str:
        return []
    import html as html_mod
    matches = re.findall(
        r"src=[\"']([^\"']+\.(?:jpg|jpeg|png|gif|webp)[^\"']*)[\"']",
        html_str, re.IGNORECASE
    )
    seen = set()
    urls = []
    for u in matches:
        u = html_mod.unescape(u)
        if u.startswith('http') and u not in seen:
            seen.add(u)
            urls.append(u)
    return urls


def export_xlsx(xlsx_bytes: bytes, saved_data: dict) -> bytes:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    cl_col, cm_col = 90, 91

    # Build row map for fast lookup
    row_map = {}
    for row in ws.iter_rows(min_row=2):
        prod_no = str(row[0].value or '').strip()
        if prod_no:
            row_map[prod_no] = row[0].row

    for prod_no, info in saved_data.items():
        key = str(prod_no).strip()
        row_idx = row_map.get(key)
        if row_idx:
            if info.get('cl_url'):
                ws.cell(row=row_idx, column=cl_col, value=info['cl_url'])
            if info.get('cm_url'):
                ws.cell(row=row_idx, column=cm_col, value=info['cm_url'])

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return buf.getvalue()


def fetch_image_as_b64(url: str) -> str:
    """이미지 URL을 base64 data URL로 반환 (CORS 프록시 대체)."""
    encoded_url = urllib.parse.quote(url, safe=':/?=&%#+@!')
    attempts = [
        {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Referer': encoded_url,
            'Accept': 'image/webp,image/apng,image/*,*/*;q=0.8',
        },
        {
            'User-Agent': 'Mozilla/5.0',
            'Referer': 'https://img.shopling.co.kr/',
            'Accept': '*/*',
        },
        {
            'User-Agent': 'Mozilla/5.0',
            'Accept': '*/*',
        },
    ]

    last_exc = None
    for headers in attempts:
        req = urllib.request.Request(encoded_url, headers=headers)
        try:
            with urllib.request.urlopen(req, timeout=20) as resp:
                data = resp.read()
                content_type = resp.headers.get('Content-Type', 'image/jpeg').split(';')[0].strip()
            b64 = base64.b64encode(data).decode('ascii')
            return f'data:{content_type};base64,{b64}'
        except Exception as exc:
            last_exc = exc

    raise RuntimeError(f'이미지 로드 실패: {url} → {last_exc}') from last_exc


def _get_drive_service():
    try:
        from googleapiclient.discovery import build
    except ImportError as exc:
        raise DriveUploadError('Missing package: pip install google-api-python-client') from exc

    # Streamlit secrets or env var (cloud)
    oauth_json = None
    try:
        oauth_json = st.secrets.get('GOOGLE_OAUTH_TOKEN_JSON')
    except Exception:
        pass
    if not oauth_json:
        oauth_json = os.environ.get('GOOGLE_OAUTH_TOKEN_JSON')

    if oauth_json:
        try:
            from google.oauth2.credentials import Credentials
            from google.auth.transport.requests import Request
            creds = Credentials.from_authorized_user_info(json.loads(oauth_json), SCOPES)
            if not creds.valid:
                if creds.expired and creds.refresh_token:
                    creds.refresh(Request())
                else:
                    raise DriveUploadError('OAuth 토큰이 만료되었습니다. 로컬에서 재인증 후 token.json을 secrets에 등록하세요.')
            return build('drive', 'v3', credentials=creds, cache_discovery=False)
        except DriveUploadError:
            raise
        except Exception as exc:
            raise DriveUploadError(f'OAuth 인증 실패: {exc}') from exc

    # Local: token.json + credentials.json
    OAUTH_CALLBACK_PORT = 60893
    OAUTH_REDIRECT_URI = f'http://localhost:{OAUTH_CALLBACK_PORT}/'
    try:
        from google.oauth2.credentials import Credentials
        from google_auth_oauthlib.flow import InstalledAppFlow
        from google.auth.transport.requests import Request
    except ImportError as exc:
        raise DriveUploadError('Missing package: pip install google-auth-oauthlib') from exc

    creds = None
    token_path = os.path.join(os.path.dirname(__file__), 'token.json')
    client_path = os.path.join(os.path.dirname(__file__), 'credentials.json')

    if os.path.exists(token_path):
        creds = Credentials.from_authorized_user_file(token_path, SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            if not os.path.exists(client_path):
                raise DriveUploadError(
                    'credentials.json 파일이 없습니다.\n'
                    'Google Cloud Console에서 OAuth 2.0 클라이언트 ID를 발급받아\n'
                    f'{client_path} 로 저장하세요.'
                )
            flow = InstalledAppFlow.from_client_secrets_file(client_path, SCOPES)
            creds = flow.run_local_server(
                port=OAUTH_CALLBACK_PORT,
                authorization_prompt_message='브라우저에서 Google 인증을 완료하세요: {url}',
                success_message='인증 완료. 이 창을 닫아도 됩니다.',
            )
        with open(token_path, 'w', encoding='utf-8') as f:
            f.write(creds.to_json())

    return build('drive', 'v3', credentials=creds, cache_discovery=False)


def _get_folder_id() -> str:
    try:
        folder_id = st.secrets.get('DRIVE_FOLDER_ID')
        if folder_id:
            return folder_id
    except Exception:
        pass
    return DRIVE_FOLDER_ID


def upload_to_drive(filename: str, data_url: str) -> dict:
    from googleapiclient.http import MediaIoBaseUpload

    file_bytes, mime_type = _parse_data_url(data_url)
    service = _get_drive_service()
    folder_id = _get_folder_id()

    escaped_name = filename.replace("'", "\\'")
    query = f"name = '{escaped_name}' and '{folder_id}' in parents and trashed = false"
    listed = service.files().list(
        q=query, spaces='drive',
        fields='files(id,name,webViewLink,modifiedTime)',
        orderBy='modifiedTime desc', pageSize=10,
    ).execute()
    existing = listed.get('files', [])

    media = MediaIoBaseUpload(io.BytesIO(file_bytes), mimetype=mime_type, resumable=False)
    if existing:
        file_id = existing[0]['id']
        item = service.files().update(
            fileId=file_id,
            body={'name': filename},
            media_body=media,
            fields='id,name,webViewLink',
        ).execute()
    else:
        metadata = {'name': filename, 'parents': [folder_id]}
        item = service.files().create(
            body=metadata, media_body=media, fields='id,name,webViewLink'
        ).execute()
        file_id = item['id']
        service.permissions().create(
            fileId=file_id,
            body={'type': 'anyone', 'role': 'reader'},
            fields='id',
        ).execute()

    fid = item['id']
    return {
        'file_id': fid,
        'name': item.get('name', filename),
        'public_url': f'https://drive.google.com/uc?export=view&id={fid}',
        'download_url': f'https://drive.google.com/uc?export=download&id={fid}',
        'web_view_url': item.get('webViewLink', ''),
    }


def delete_drive_file(file_id: str):
    service = _get_drive_service()
    try:
        service.files().delete(fileId=file_id).execute()
    except Exception as exc:
        raise DriveUploadError(f'Drive 파일 삭제 실패: {exc}') from exc


def update_drive_file(file_id: str, filename: str, data_url: str) -> dict:
    from googleapiclient.http import MediaIoBaseUpload

    file_bytes, mime_type = _parse_data_url(data_url)
    service = _get_drive_service()
    media = MediaIoBaseUpload(io.BytesIO(file_bytes), mimetype=mime_type, resumable=False)
    item = service.files().update(
        fileId=file_id,
        body={'name': filename} if filename else None,
        media_body=media,
        fields='id,name,webViewLink',
    ).execute()
    fid = item['id']
    return {
        'file_id': fid,
        'name': item.get('name', filename or ''),
        'public_url': f'https://drive.google.com/uc?export=view&id={fid}',
        'download_url': f'https://drive.google.com/uc?export=download&id={fid}',
        'web_view_url': item.get('webViewLink', ''),
    }


def _parse_data_url(data_url: str):
    if not data_url.startswith('data:') or ';base64,' not in data_url:
        raise DriveUploadError('data_url 형식이 올바르지 않습니다')
    header, b64_data = data_url.split(';base64,', 1)
    mime_type = header[5:] if header.startswith('data:') else 'application/octet-stream'
    try:
        return base64.b64decode(b64_data), mime_type
    except Exception as exc:
        raise DriveUploadError(f'Base64 디코드 실패: {exc}') from exc
