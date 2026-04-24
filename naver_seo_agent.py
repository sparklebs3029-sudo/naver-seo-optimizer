#!/usr/bin/env python3
"""
네이버 SEO 상품명 최적화 + 상품 소싱 에이전트
"""

import os
import re
import json
import time
import argparse
from datetime import datetime, timedelta
import requests
import google.generativeai as genai
import openpyxl


# ── 네이버 쇼핑 카테고리 ────────────────────────────────────────────
NAVER_CATEGORIES = {
    "패션의류":     "50000000",
    "패션잡화":     "50000001",
    "화장품/미용":  "50000002",
    "디지털/가전":  "50000003",
    "가구/인테리어":"50000004",
    "출산/육아":    "50000005",
    "식품":         "50000006",
    "스포츠/레저":  "50000007",
    "생활/건강":    "50000008",
    "자동차용품":   "50000011",
    "완구/취미":    "50000013",
    "문구/오피스":  "50000014",
    "반려동물용품": "50000015",
    "농수축산물":   "50000016",
}

# ── 배송 관련 금지 문구 ─────────────────────────────────────────────
DELIVERY_TERMS = [
    "오늘출발", "당일배송", "빠른배송", "무료배송", "익일배송",
    "새벽배송", "로켓배송", "총알배송", "당일출발", "빠른출발",
]

# ── 홍보성/주관적 금지 수식어 ───────────────────────────────────────
PROMO_WORDS = [
    "최고", "최저가", "전국최저가", "특가", "최저", "강추", "완전", "대박",
    "개꿀", "인기폭발", "베스트", "핫딜", "할인", "세일", "역대급", "초특가",
    "가성비최고", "품질보장", "정품보장", "100%정품",
]

# ── 브랜드명 목록 (원본에 없으면 제거) ──────────────────────────────
BRAND_NAMES = [
    # 글로벌 스포츠
    "나이키", "아디다스", "뉴발란스", "퓨마", "리복", "컨버스", "반스",
    "언더아머", "챔피언", "휠라",
    # 아웃도어
    "노스페이스", "블랙야크", "K2", "코오롱스포츠", "파타고니아", "아이더",
    "밀레", "네파", "몽벨", "콜롬비아",
    # 국내·글로벌 SPA
    "유니클로", "에잇세컨즈", "탑텐", "스파오", "지오다노", "자라", "H&M", "망고", "갭",
    # 국내·글로벌 패션
    "빈폴", "헤지스", "라코스테", "폴로", "타미힐피거", "캘빈클라인",
    "리바이스", "게스", "MLB", "이랜드",
    # 명품
    "구찌", "루이비통", "샤넬", "버버리", "프라다", "MCM", "코치",
    "마이클코어스", "발렌시아가", "생로랑",
    # 잠옷·홈웨어·이너웨어
    "편한밤", "비비안", "비너스", "빅토리아시크릿", "캘빈클라인", "에스콰이아",
    "트라이", "쌍방울", "보디가드", "예스민", "에스마인", "슈크림",
    # 스포츠·레저 추가
    "안다르", "젝시믹스", "뮬라웨어", "룰루레몬",
]

# ── 소싱 금지 품목 그룹 ─────────────────────────────────────────────
PROHIBITED_GROUPS = {
    "총기/무기류":   ["총기", "권총", "소총", "공기총", "BB탄총", "도검", "폭발물", "폭탄", "화약", "탄약", "수류탄"],
    "마약/향정신성": ["마약", "대마", "필로폰", "코카인", "헤로인", "LSD", "히로뽕", "향정신성"],
    "유해화학물질":  ["독극물", "청산가리", "시안화", "화학무기"],
    "성인용품":      ["성인용품", "음란", "성인토이"],
    "전문의약품":    ["처방전", "전문의약품", "향정신성의약품"],
}

# ── 카테고리별 시드 키워드 ───────────────────────────────────────────
CATEGORY_SEED_KEYWORDS: dict[str, list[str]] = {
    "패션의류":     ["티셔츠", "청바지", "원피스", "코트", "패딩", "니트", "블라우스", "바지", "자켓", "후드티"],
    "패션잡화":     ["가방", "지갑", "벨트", "모자", "선글라스", "시계", "스카프", "신발", "슬리퍼", "백팩"],
    "화장품/미용":  ["마스크팩", "에센스", "크림", "선크림", "립스틱", "파운데이션", "샴푸", "바디로션", "향수", "세럼"],
    "디지털/가전":  ["이어폰", "충전기", "보조배터리", "블루투스스피커", "스마트워치", "마우스", "키보드", "웹캠", "USB허브", "케이블"],
    "가구/인테리어": ["책상", "의자", "침대", "소파", "수납장", "조명", "커튼", "카펫", "선반", "행거"],
    "출산/육아":    ["기저귀", "분유", "유모차", "아기띠", "장난감", "이유식", "아기옷", "젖병", "카시트", "아기침대"],
    "식품":         ["과자", "음료", "라면", "커피", "건강식품", "견과류", "초콜릿", "홍삼", "프로틴", "비타민"],
    "스포츠/레저":  ["운동화", "레깅스", "요가매트", "덤벨", "등산화", "텐트", "낚시", "골프장갑", "수영복", "자전거"],
    "생활/건강":    ["칫솔", "세제", "주방용품", "멀티탭", "마스크", "비타민", "체중계", "공기청정기", "청소기", "욕실용품"],
    "자동차용품":   ["방향제", "블랙박스", "차량충전기", "세차용품", "차량매트", "햇빛가리개", "네비게이션", "타이어", "차량배터리", "열선"],
    "완구/취미":    ["레고", "보드게임", "피규어", "퍼즐", "드론", "RC카", "프라모델", "색연필", "그림도구", "미니어처"],
    "문구/오피스":  ["볼펜", "노트", "스케줄러", "테이프", "가위", "포스트잇", "계산기", "파일", "스탬프", "화이트보드"],
    "반려동물용품": ["사료", "간식", "패드", "장난감", "리드줄", "하네스", "하우스", "캐리어", "영양제", "샴푸"],
    "농수축산물":   ["쌀", "김치", "고구마", "감자", "양파", "마늘", "사과", "배", "한우", "굴비"],
}

# ── 시스템 프롬프트 ─────────────────────────────────────────────────
GEMINI_CONFIG = {"temperature": 0}

KEYWORD_SYSTEM = """당신은 네이버 쇼핑 검색 전문가입니다.
상품명을 보고 구매자들이 실제로 네이버에서 검색할 키워드 후보를 생성합니다."""

CLASSIFY_SYSTEM = """당신은 네이버 쇼핑 SEO 전문가입니다.
주어진 키워드 목록을 핵심 키워드와 보조 단어로 분류하고 반드시 JSON으로만 응답합니다."""

OPTIMIZE_SYSTEM = """당신은 네이버 스마트스토어 SEO 전문가입니다.
네이버 검색 알고리즘과 쇼핑 검색 최적화에 깊은 이해를 가지고 있습니다.

네이버 SEO 상품명 최적화 원칙:
1. 핵심 키워드를 앞쪽에 배치 (검색 노출 우선순위)
2. 상품의 주 구매 타겟(예: 자취생, 부모님선물, 캠핑족, 신혼부부 등)이나 핵심 사용 상황을 파악하여 상품명 가장 앞에 1~2단어로 자연스럽게 배치
3. 브랜드명 + 카테고리 + 세부 특성 + 타겟/용도 조합
4. 제공된 고검색량 키워드를 자연스럽게 포함
5. 반드시 공백 포함 25자 이상 50자 이하 (네이버 SEO 최적 범위)
6. 특수문자 일절 사용 금지 (**, [], ##, ~~, /, | 등 모든 기호 제외)
7. 배송 관련 문구 절대 포함 금지 (오늘출발, 당일배송, 빠른배송, 무료배송 등)
8. 홍보성·주관적 수식어 사용 금지 (최고, 특가, 강추, 대박, 베스트 등)
9. 답변은 반드시 순수 텍스트 상품명만 출력, 마크다운 서식 사용 금지
10. 원본에 없는 소재·형태·디자인 속성 추가 절대 금지
    — 소재: 면, 폴리, 실크, 레이스, 니트, 데님, 린넨 등
    — 형태: 긴팔, 반팔, 민소매, 긴바지, 반바지 등
    — 디자인: 레이스, 프릴, 리본, 체크, 스트라이프 등
    원본에 명시된 속성만 유지하고, 확인되지 않은 속성은 절대 추가하지 않는다."""

VERIFY_SYSTEM = """당신은 네이버 스마트스토어 상품명 검수 전문가입니다.
최적화된 상품명이 품질 기준을 충족하는지 검토하고 필요시 수정합니다.
⭐ DataLab 검증 키워드(붙여쓰기 복합어 포함)는 절대 분해하거나 삭제하지 마세요.
원본에 없는 소재(레이스, 면, 실크 등), 형태(긴팔, 반팔 등), 디자인(프릴, 리본 등) 단어가
DataLab 검증 키워드가 아닌 형태로 단독 추가된 경우에만 제거하세요."""


# ── 설정 파일 로드 ──────────────────────────────────────────────────
def load_config() -> dict[str, str]:
    config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.txt")
    config = {}
    if not os.path.exists(config_path):
        return config
    with open(config_path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            value = value.strip()
            if value and not value.startswith("여기에"):
                config[key.strip()] = value
    return config


# ── Naver API 재시도 헬퍼 ────────────────────────────────────────────
def _post_with_retry(url: str, headers: dict, body: dict, max_retries: int = 3) -> dict:
    """POST 요청을 최대 3회 재시도합니다. 호출 사이 0.2초 대기."""
    last_err = None
    for attempt in range(max_retries):
        if attempt > 0:
            time.sleep(0.2 * attempt)
        try:
            resp = requests.post(url, headers=headers, json=body, timeout=10)
            resp.raise_for_status()
            return resp.json()
        except requests.exceptions.HTTPError as e:
            last_err = e
            if e.response is not None and e.response.status_code < 500:
                raise
        except requests.exceptions.RequestException as e:
            last_err = e
    raise last_err


# ── 상품코드 제거 ──────────────────────────────────────────────────
def strip_product_code(name: str) -> str:
    """상품명 끝의 상품코드(예: LKO550, SEG-1008129, UJS-065)를 제거합니다."""
    return re.sub(r'\s+[A-Z]{1,6}[-]?\d{2,}[-]?\d*$', '', name).strip()


# ── 단어 풀 구성 / 필터 ────────────────────────────────────────────
def build_word_pool(original: str, top_keywords: list[str]) -> set[str]:
    """원본 단어 + 트렌드 키워드 단어로 허용 풀을 구성합니다.
    인접 2-gram도 포함해 복합어(잠옷+원피스→잠옷원피스) 허용."""
    pool: set[str] = set()
    orig_tokens = original.split()
    pool.update(orig_tokens)
    for i in range(len(orig_tokens) - 1):
        pool.add(orig_tokens[i] + orig_tokens[i + 1])
    for kw in top_keywords:
        pool.add(kw)
        kw_tokens = kw.split()
        pool.update(kw_tokens)
        for i in range(len(kw_tokens) - 1):
            pool.add(kw_tokens[i] + kw_tokens[i + 1])
    return pool


def filter_to_pool(name: str, pool: set[str]) -> str:
    """pool에 없는 단어를 제거합니다."""
    kept = [w for w in name.split() if w in pool]
    return re.sub(r'\s+', ' ', ' '.join(kept)).strip()


# ── 1단계: 키워드 후보 에이전트 ────────────────────────────────────
def generate_keyword_candidates(
    original: str,
    model: "genai.GenerativeModel",
    feedback: str = "",
) -> list[str]:
    feedback_section = ""
    if feedback:
        feedback_section = (
            f"⚠️ 이전 시도 실패 이유 (이를 참고해 더 적절한 키워드를 생성하세요):\n{feedback}\n\n"
        )
    prompt = (
        f"{feedback_section}"
        "다음 상품명과 관련된 네이버 쇼핑 검색 키워드 후보를 15개 생성해주세요.\n\n"
        "네이버 사용자들은 두 가지 방식으로 검색합니다:\n"
        "① 띄어쓰기 형태: '무지 에코백', '가로형 에코백'\n"
        "② 붙여쓰기 복합 키워드: '무지에코백가로형', '캔버스에코백', '여행용숄더백'\n\n"
        "⚠️ 붙여쓰기 복합 키워드(②)가 실제 네이버 검색량의 대부분을 차지합니다.\n"
        "반드시 15개 중 절반 이상(8개 이상)은 붙여쓰기 복합 키워드로 생성하세요.\n\n"
        "규칙:\n"
        "- 2~4개 단어를 조합한 구체적인 키워드\n"
        "- 브랜드명, 배송 관련 단어는 제외\n"
        "- 상품의 용도, 소재, 형태, 타겟을 조합\n\n"
        "JSON 배열 형식으로만 답변하세요: [\"키워드1\", \"키워드2\", ...]\n\n"
        f"상품명: {original}"
    )
    try:
        response = model.generate_content(prompt)
        raw = response.text.strip()
        if raw.startswith("```"):
            lines = raw.splitlines()
            raw = "\n".join(lines[1:-1] if lines[-1].strip() == "```" else lines[1:])
        # JSON 배열이 텍스트 안에 섞여 있는 경우 추출
        m = re.search(r'\[[\s\S]*?\]', raw)
        if m:
            raw = m.group(0)
        result = json.loads(raw)
        if isinstance(result, list) and result:
            return result
    except Exception:
        pass
    # 폴백: 원본 단어를 키워드로 사용
    return [w for w in original.split() if len(w) >= 2][:15] or [original]


def detect_category(original: str, model: genai.GenerativeModel) -> str:
    categories_list = "\n".join(f"- {name}" for name in NAVER_CATEGORIES)
    prompt = (
        "다음 상품명이 네이버 쇼핑의 어떤 카테고리에 속하는지 판단하세요.\n"
        "아래 목록 중 가장 적합한 카테고리명 하나만 정확히 답변하세요:\n\n"
        f"{categories_list}\n\n"
        f"상품명: {original}"
    )
    try:
        response = model.generate_content(prompt)
        category_name = response.text.strip()
        for name in NAVER_CATEGORIES:
            if name in category_name:
                return NAVER_CATEGORIES[name]
        return NAVER_CATEGORIES.get(category_name, "50000000")
    except Exception:
        # API 실패 시 상품명 키워드로 간단 추정
        _fashion = ["원피스", "티셔츠", "바지", "치마", "코트", "자켓", "블라우스", "니트", "패딩", "후드"]
        if any(w in original for w in _fashion):
            return "50000000"  # 패션의류
        _bag = ["가방", "지갑", "모자", "신발", "슬리퍼", "샌들", "부츠"]
        if any(w in original for w in _bag):
            return "50000001"  # 패션잡화
        return "50000000"  # 기본값: 패션의류


# ── 2단계: 데이터랩 조회 ────────────────────────────────────────────
def query_search_trend(keywords: list[str], client_id: str, client_secret: str) -> dict[str, float]:
    end_date   = datetime.now()
    start_date = end_date - timedelta(days=90)
    headers = {
        "X-Naver-Client-Id":     client_id,
        "X-Naver-Client-Secret": client_secret,
        "Content-Type":          "application/json",
    }
    results = {}
    for i in range(0, len(keywords), 5):
        batch = keywords[i:i + 5]
        body  = {
            "startDate":     start_date.strftime("%Y-%m-%d"),
            "endDate":       end_date.strftime("%Y-%m-%d"),
            "timeUnit":      "month",
            "keywordGroups": [{"groupName": kw, "keywords": [kw]} for kw in batch],
        }
        try:
            data = _post_with_retry("https://openapi.naver.com/v1/datalab/search", headers, body)
            for result in data.get("results", []):
                ratios = [p["ratio"] for p in result.get("data", [])]
                results[result["title"]] = sum(ratios) / len(ratios) if ratios else 0.0
        except Exception:
            pass  # DataLab 실패 시 해당 배치 건너뜀
        time.sleep(0.2)
    return results


def query_shopping_insight(keywords: list[str], category_id: str, client_id: str, client_secret: str) -> dict[str, float]:
    end_date   = datetime.now()
    start_date = end_date - timedelta(days=90)
    headers = {
        "X-Naver-Client-Id":     client_id,
        "X-Naver-Client-Secret": client_secret,
        "Content-Type":          "application/json",
    }
    results = {}
    for i in range(0, len(keywords), 5):
        batch = keywords[i:i + 5]
        body  = {
            "startDate": start_date.strftime("%Y-%m-%d"),
            "endDate":   end_date.strftime("%Y-%m-%d"),
            "timeUnit":  "month",
            "category":  category_id,
            "keyword":   [{"name": kw, "param": [kw]} for kw in batch],
        }
        try:
            data = _post_with_retry(
                "https://openapi.naver.com/v1/datalab/shopping/category/keywords", headers, body
            )
            for result in data.get("results", []):
                ratios = [p["ratio"] for p in result.get("data", [])]
                results[result["title"]] = sum(ratios) / len(ratios) if ratios else 0.0
        except Exception:
            pass  # DataLab 실패 시 해당 배치 건너뜀
        time.sleep(0.2)
    return results


def combine_and_select(
    search_scores:   dict[str, float],
    shopping_scores: dict[str, float],
    keywords:        list[str],
    n:               int = 5,
) -> list[str]:
    combined = {
        kw: search_scores.get(kw, 0.0) * 0.4 + shopping_scores.get(kw, 0.0) * 0.6
        for kw in keywords
    }
    return [kw for kw, _ in sorted(combined.items(), key=lambda x: x[1], reverse=True)[:n]]


# ── 2.5단계: 키워드 분류 (핵심 / 보조) ────────────────────────────
def classify_keywords(
    top_keywords: list[str],
    original: str,
    model: "genai.GenerativeModel",
) -> tuple[list[str], list[str]]:
    """상위 키워드를 핵심 키워드(3개)와 보조 단어(최대 3개)로 분류한다.

    핵심 키워드: 상품을 직접 설명하는 복합 키워드. 상품명에서 쪼개지 않고 사용.
    보조 단어: 핵심 키워드 바로 뒤에 붙였을 때 역방향으로 읽으면 유효한 검색어가 되는 짧은 단어.
    """
    prompt = (
        f"상품명: {original}\n"
        f"인기 키워드 목록: {', '.join(top_keywords)}\n\n"
        "아래 기준으로 분류하세요.\n\n"
        "핵심 키워드 3개: 검색량이 높고 상품을 직접 설명하는 복합 키워드. 상품명에서 절대 쪼개지 않음.\n"
        "보조 단어 최대 3개: 핵심 키워드 바로 뒤에 붙였을 때 역방향으로 읽으면 유효한 검색어가 되는 짧은 단어(1~3음절).\n"
        "  예: 핵심='반팔 롱 원피스', 보조='여자' → '반팔 롱 원피스 여자'를 뒤에서 읽으면 '여자원피스'\n\n"
        "핵심 키워드가 3개 미만이면 관련 키워드를 조합해 보완하세요.\n"
        '다음 JSON 형식으로만 답변: {"core": ["핵심1", "핵심2", "핵심3"], "aux": ["보조1", "보조2", "보조3"]}'
    )
    try:
        response = model.generate_content(prompt)
        raw = response.text.strip()
        if raw.startswith("```"):
            lines = raw.splitlines()
            raw = "\n".join(lines[1:-1] if lines[-1].strip() == "```" else lines[1:])
        data = json.loads(raw)
        core = data.get("core", top_keywords[:3])[:3]
        aux  = data.get("aux", [])[:3]
        return core, aux
    except Exception:
        return top_keywords[:3], []


def build_guide_name(core_keywords: list[str], aux_words: list[str]) -> str:
    """핵심 키워드 사이에 보조 단어를 삽입해 역순 조합 구조의 상품명을 만든다.

    보조 단어를 최대한 포함하면서 50자 이하가 되도록 뒤쪽 보조 단어부터 제거한다.
    """
    for num_aux in range(len(aux_words), -1, -1):
        parts: list[str] = []
        for i, core in enumerate(core_keywords):
            parts.append(core)
            if i < num_aux and i < len(aux_words) and aux_words[i]:
                parts.append(aux_words[i])
        name = " ".join(parts).strip()
        if len(name) <= 50:
            return name
    name = " ".join(core_keywords)
    return name[:50].rsplit(" ", 1)[0] if len(name) > 50 else name


# ── 3단계: 최적화 에이전트 ─────────────────────────────────────────
def optimize_name(
    original: str,
    core_keywords: list[str],
    aux_words: list[str],
    model: "genai.GenerativeModel",
) -> str:
    """가이드 역순 조합 구조로 상품명을 최적화한다.

    사이즈·1+1 정보가 없으면 알고리즘으로 구성한 이름을 그대로 반환한다.
    해당 정보가 있으면 AI가 삽입 위치를 결정한다.
    """
    guide_name = build_guide_name(core_keywords, aux_words)

    has_size  = bool(re.search(r'\d+\s*(cm|mm|m|L|ml|g|kg|인치|평|구|포|매|개|장|켤레|족)', original, re.IGNORECASE))
    has_bonus = bool(re.search(r'1\+1|2\+1|증정|사은품', original))

    if not has_size and not has_bonus and 25 <= len(guide_name) <= 50:
        return guide_name

    size_note  = "원본에 사이즈/규격 정보가 있으므로 반드시 유지하고 상품명 뒤쪽에 배치하세요." if has_size  else "원본에 사이즈 정보가 없으므로 임의로 추가하지 마세요."
    bonus_note = "원본에 1+1/증정 정보가 있으므로 반드시 유지하고 핵심키워드 뒤에 배치하세요." if has_bonus else "원본에 1+1/증정 정보가 없으므로 임의로 추가하지 마세요."
    core_str   = " / ".join(core_keywords)
    aux_str    = " / ".join(aux_words) if aux_words else "없음"

    prompt = (
        "아래 구조를 따라 네이버 SEO 상품명을 완성하세요.\n\n"
        "▶ 역순 조합 구조 (이 순서를 반드시 유지):\n"
        f"  핵심 키워드: {core_str}\n"
        f"  보조 단어: {aux_str}\n"
        "  조합 규칙: [핵심1] [보조1] [핵심2] [보조2] [핵심3] [보조3]\n"
        "  — 핵심 키워드는 절대 쪼개지 말 것. 보조 단어는 핵심 키워드 바로 뒤에 삽입.\n\n"
        f"▶ 사이즈/규격: {size_note}\n"
        f"▶ 1+1/증정: {bonus_note}\n"
        "▶ 반드시 공백 포함 25자 이상 50자 이하\n"
        "▶ 특수문자·배송 문구·홍보 수식어·원본에 없는 브랜드명 금지\n"
        "▶ 순수 텍스트 상품명 1개만 출력\n\n"
        f"원본 상품명: {original}\n"
        f"초안(참고용): {guide_name}"
    )
    try:
        response = model.generate_content(prompt)
        return response.text.strip()
    except Exception:
        return guide_name  # AI 호출 실패 시 알고리즘 결과 반환


# ── 4단계: 검수 (코드 규칙 + AI) ───────────────────────────────────
def _remove_duplicate_words(text: str) -> str:
    seen:   set[str]  = set()
    result: list[str] = []
    for word in text.split():
        if word not in seen:
            seen.add(word)
            result.append(word)
    return " ".join(result)


def clean_by_rules(name: str, original: str = "") -> str:
    name = re.sub(r'[*#~`\[\](){}|/\\]', '', name)
    for term in DELIVERY_TERMS:
        name = name.replace(term, '')
    for word in PROMO_WORDS:
        name = re.sub(rf'\b{re.escape(word)}\b', '', name)
    for brand in BRAND_NAMES:
        if brand in name and brand not in original:
            name = name.replace(brand, '')
    name = _remove_duplicate_words(name)
    name = re.sub(r'\s+', ' ', name).strip()
    if len(name) > 50:
        name = name[:50].rsplit(' ', 1)[0].strip()
    return name


def enforce_min_length(name: str, original: str, top_keywords: list[str], model: genai.GenerativeModel) -> str:
    """25자 미만인 경우 키워드를 추가해 재확장합니다."""
    if len(name) >= 25:
        return name
    keywords_str = ", ".join(top_keywords)
    prompt = (
        f"다음 상품명이 {len(name)}자로 너무 짧습니다. 반드시 25자 이상 50자 이하로 늘려주세요.\n"
        f"원본 상품명: {original}\n"
        f"현재 상품명: {name}\n"
        f"참고 키워드: {keywords_str}\n"
        "현재 상품명을 기반으로 관련 키워드를 자연스럽게 추가해 25자 이상으로 확장하세요.\n"
        "특수문자, 배송 문구, 홍보 수식어는 사용하지 마세요.\n"
        "순수 텍스트 상품명만 답변하세요."
    )
    result = clean_by_rules(model.generate_content(prompt).text.strip(), original)
    return result if len(result) >= 25 else name


def verify_name(
    original: str,
    optimized: str,
    model: genai.GenerativeModel,
    allowed_keywords: list[str] | None = None,
) -> tuple[str, str | None]:
    length = len(optimized)
    length_note = ""
    if length < 25:
        length_note = f"현재 {length}자로 너무 짧습니다. 관련 키워드를 추가해 25자 이상으로 늘려주세요."
    elif length > 50:
        length_note = f"현재 {length}자로 너무 깁니다. 50자 이하로 줄여주세요."

    allowed_section = ""
    if allowed_keywords:
        kw_str = ", ".join(allowed_keywords[:10])
        allowed_section = (
            f"\n⭐ 네이버 DataLab 검증 키워드 (절대 제거 금지): {kw_str}\n"
            "   — 위 키워드는 원본에 없더라도 반드시 유지하세요.\n"
        )

    prompt = (
        "아래 최적화된 상품명을 검수하고 필요시 수정해주세요.\n\n"
        f"원본 상품명: {original}\n"
        f"최적화된 상품명: {optimized}\n"
        f"{('⚠️ 글자수 조정 필요: ' + length_note) if length_note else ''}"
        f"{allowed_section}\n"
        "검수 기준:\n"
        "1. 의미 없는 수식어 제거 (최고, 대박, 완전, 특가, 강추 등)\n"
        "2. 중복 단어 제거\n"
        "3. 원본 상품명에 없는 무관한 브랜드명 제거 (원본에 있는 브랜드는 유지)\n"
        "4. 원본 상품과 완전히 무관한 키워드 제거 (DataLab 검증 키워드는 제외)\n"
        "5. 네이버 금지 표현 제거\n"
        "6. 공백 포함 25자 이상 50자 이하 유지 — 절대 25자 미만으로 줄이지 말 것 (단어 삭제 시 글자 수 확인 필수)\n"
        "7. 원본에 없는 소재·형태·디자인 속성은 제거하되, DataLab 검증 키워드는 예외\n"
        "   — DataLab 검증 키워드에 속성어가 포함된 복합어(예: 면원피스, 니트원피스, 린넨비치웨어)는 원본에 해당 속성어가 없더라도 반드시 유지\n"
        "   — DataLab 검증 키워드가 아닌 단독 속성어 예: 긴팔, 반팔, 레이스, 면, 실크, 니트 등\n\n"
        "다음 JSON 형식으로만 답변하세요:\n"
        '{"final_name": "최종 상품명", "issues": "수정 사항 설명 (없으면 null)"}'
    )
    try:
        response = model.generate_content(prompt)
        raw = response.text.strip()
        if raw.startswith("```"):
            lines = raw.splitlines()
            raw = "\n".join(lines[1:-1] if lines[-1].strip() == "```" else lines[1:])
        data = json.loads(raw)
        final_name = data.get("final_name") or optimized
        issues_raw = data.get("issues")
        issues = issues_raw if isinstance(issues_raw, str) and issues_raw.lower() not in ("null", "없음", "none", "") else None
        return final_name, issues
    except Exception:
        return optimized, None


# ── 소싱 에이전트 ───────────────────────────────────────────────────
def search_naver_shopping(keyword: str, client_id: str, client_secret: str, display: int = 5) -> list[dict]:
    """네이버 쇼핑 검색 API로 인기 상품을 검색합니다."""
    resp = requests.get(
        "https://openapi.naver.com/v1/search/shop.json",
        headers={
            "X-Naver-Client-Id":     client_id,
            "X-Naver-Client-Secret": client_secret,
        },
        params={"query": keyword, "display": display, "sort": "sim"},
        timeout=10,
    )
    resp.raise_for_status()
    items = resp.json().get("items", [])
    for item in items:
        item["title"] = re.sub(r"<[^>]+>", "", item.get("title", ""))
    return items


def extract_search_query(original: str, model) -> str:
    """원본명에서 핵심 카테고리 키워드 2~3개만 추출해 검색 쿼리로 반환합니다."""
    prompt = (
        f"상품명: {original}\n\n"
        "네이버 쇼핑 검색에 사용할 핵심 카테고리 키워드 2~3개만 추출하세요.\n"
        "형용사·수식어·소재명은 제외, 상품 종류를 나타내는 단어만 남기세요.\n"
        "예: '끈 조절 슬립 잠옷원피스' → '슬립 잠옷원피스'\n"
        "예: '여성 루즈핏 반팔 롱 면 원피스' → '여성 롱 원피스'\n"
        "결과만 출력하세요."
    )
    try:
        return model.generate_content(prompt).text.strip()
    except Exception:
        return original


def fallback_by_shopping_search(
    original: str,
    naver_id: str,
    naver_secret: str,
    model,
    classify_model,
) -> str:
    """3회 재시도 실패 시 네이버 쇼핑 상위 상품명을 참고해 폴백 최적화합니다."""
    search_query = extract_search_query(original, classify_model)
    items = search_naver_shopping(search_query, naver_id, naver_secret, display=5)

    competitor_names = []
    for item in items:
        name = item.get("title", "")
        for brand in BRAND_NAMES:
            name = name.replace(brand, "")
        name = name.strip()
        if name:
            competitor_names.append(name)

    if not competitor_names:
        prompt = (
            f"원본 상품명: {original}\n\n"
            "원본 상품명을 네이버 SEO에 맞게 직접 확장하세요.\n"
            "▶ 공백 포함 25자 이상 50자 이하\n"
            "▶ 소재, 형태, 타겟, 용도 키워드를 추가\n"
            "▶ 특수문자·배송 문구·홍보 수식어 금지\n"
            "▶ 순수 텍스트 상품명 1개만 출력"
        )
        try:
            result = model.generate_content(prompt).text.strip()
            cleaned = clean_by_rules(result, original)
            if len(cleaned) < 25:
                cleaned = enforce_min_length(cleaned, original, [], model)
            return cleaned
        except Exception:
            return original

    names_str = "\n".join(f"- {n}" for n in competitor_names[:5])
    prompt = (
        f"원본 상품명: {original}\n\n"
        f"네이버 쇼핑 상위 노출 유사 상품명 (참고용):\n{names_str}\n\n"
        "위 상품명들의 키워드 패턴을 참고해 원본 상품명을 네이버 SEO에 맞게 최적화하세요.\n"
        "▶ 공백 포함 25자 이상 50자 이하\n"
        "▶ 원본과 동일한 상품 유형 유지\n"
        "▶ 특수문자·배송 문구·홍보 수식어·원본에 없는 브랜드명 금지\n"
        "▶ 순수 텍스트 상품명 1개만 출력"
    )
    try:
        result = model.generate_content(prompt).text.strip()
        cleaned = clean_by_rules(result, original)
        if len(cleaned) < 25:
            cleaned = enforce_min_length(cleaned, original, competitor_names, model)
        return cleaned
    except Exception:
        return original


def is_prohibited(text: str, active_groups: list[str]) -> bool:
    """텍스트에 금지 키워드가 포함되어 있는지 확인합니다."""
    text_lower = text.lower()
    for group in active_groups:
        for kw in PROHIBITED_GROUPS.get(group, []):
            if kw in text_lower:
                return True
    return False


def get_trending_products(
    category_name:      str,
    category_id:        str,
    period_days:        int,
    client_id:          str,
    client_secret:      str,
    active_prohibited:  list[str],
    extra_prohibited:   list[str],
    top_n:              int = 5,
) -> list[dict]:
    """트렌딩 키워드 조회 후 상품 소싱 결과를 반환합니다."""
    seed_keywords = CATEGORY_SEED_KEYWORDS.get(category_name, [])
    if not seed_keywords:
        return []

    end_date   = datetime.now()
    start_date = end_date - timedelta(days=period_days)
    headers = {
        "X-Naver-Client-Id":     client_id,
        "X-Naver-Client-Secret": client_secret,
        "Content-Type":          "application/json",
    }

    # DataLab 쇼핑인사이트로 시드 키워드 트렌드 점수 조회
    trend_scores: dict[str, float] = {}
    for i in range(0, len(seed_keywords), 5):
        batch = seed_keywords[i:i + 5]
        body  = {
            "startDate": start_date.strftime("%Y-%m-%d"),
            "endDate":   end_date.strftime("%Y-%m-%d"),
            "timeUnit":  "date",
            "category":  category_id,
            "keyword":   [{"name": kw, "param": [kw]} for kw in batch],
        }
        try:
            data = _post_with_retry(
                "https://openapi.naver.com/v1/datalab/shopping/category/keywords", headers, body
            )
            for result in data.get("results", []):
                ratios = [p["ratio"] for p in result.get("data", [])]
                trend_scores[result["title"]] = sum(ratios) / len(ratios) if ratios else 0.0
        except Exception:
            pass
        time.sleep(0.2)

    # 상위 N개 트렌딩 키워드
    top_keywords = sorted(trend_scores.items(), key=lambda x: x[1], reverse=True)[:top_n]

    # 각 키워드별 쇼핑 검색 + 금지 필터 적용
    results: list[dict] = []
    for keyword, score in top_keywords:
        try:
            products = search_naver_shopping(keyword, client_id, client_secret, display=5)
            for product in products:
                full_text = " ".join([
                    product.get("title", ""),
                    product.get("category1", ""),
                    product.get("category2", ""),
                    product.get("category3", ""),
                ])
                # 금지 그룹 필터
                if is_prohibited(full_text, active_prohibited):
                    continue
                # 추가 금지 키워드 필터
                if any(kw in full_text for kw in extra_prohibited if kw):
                    continue

                price = product.get("lprice", "")
                results.append({
                    "키워드":     keyword,
                    "트렌드점수": round(score, 1),
                    "상품명":     product.get("title", ""),
                    "최저가":     f"{int(price):,}원" if price else "-",
                    "쇼핑몰":     product.get("mallName", ""),
                    "카테고리":   product.get("category1", ""),
                    "링크":       product.get("link", ""),
                })
            time.sleep(0.2)
        except Exception:
            pass

    return results


# ── 출력 파일 경로 생성 ─────────────────────────────────────────────
def get_output_path(input_path: str) -> str:
    import glob
    dir_name      = os.path.dirname(os.path.abspath(input_path))
    base_name     = os.path.splitext(os.path.basename(input_path))[0]
    now           = datetime.now()
    datetime_str  = now.strftime("%Y%m%d%H%M")   # 연월일시분
    today_str     = now.strftime("%Y%m%d")        # 날짜 (번호 기준)

    # 오늘 생성된 파일 수로 번호 결정 (00시 기준 초기화)
    existing = glob.glob(os.path.join(dir_name, f"*_최적화_{today_str}*.xlsx"))
    n = len(existing) + 1

    candidate = os.path.join(dir_name, f"{base_name}_최적화_{datetime_str}_{n}.xlsx")
    while os.path.exists(candidate):
        n += 1
        candidate = os.path.join(dir_name, f"{base_name}_최적화_{datetime_str}_{n}.xlsx")
    return candidate


# ── 메인 (BAT 파일용) ───────────────────────────────────────────────
def main() -> int:
    parser = argparse.ArgumentParser(description="네이버 SEO 롱테일 상품명 최적화 에이전트")
    parser.add_argument("input", help="처리할 엑셀 파일 경로 (.xlsx)")
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"[오류] 파일을 찾을 수 없습니다: {args.input}")
        return 1

    config       = load_config()
    gemini_key   = os.environ.get("GEMINI_API_KEY")      or config.get("GEMINI_API_KEY")
    naver_id     = os.environ.get("NAVER_CLIENT_ID")     or config.get("NAVER_CLIENT_ID")
    naver_secret = os.environ.get("NAVER_CLIENT_SECRET") or config.get("NAVER_CLIENT_SECRET")

    if not gemini_key:
        print("[오류] GEMINI_API_KEY가 config.txt에 없습니다.")
        return 1
    if not naver_id or not naver_secret:
        print("[오류] NAVER_CLIENT_ID 또는 NAVER_CLIENT_SECRET이 config.txt에 없습니다.")
        return 1

    genai.configure(api_key=gemini_key)
    keyword_model  = genai.GenerativeModel("gemini-2.0-flash", system_instruction=KEYWORD_SYSTEM,  generation_config=GEMINI_CONFIG)
    classify_model = genai.GenerativeModel("gemini-2.0-flash", system_instruction=CLASSIFY_SYSTEM, generation_config=GEMINI_CONFIG)
    optimize_model = genai.GenerativeModel("gemini-2.0-flash", system_instruction=OPTIMIZE_SYSTEM, generation_config=GEMINI_CONFIG)
    verify_model   = genai.GenerativeModel("gemini-2.0-flash", system_instruction=VERIFY_SYSTEM,   generation_config=GEMINI_CONFIG)

    wb = openpyxl.load_workbook(args.input)
    ws = wb.active
    H_COL = 8

    data_rows = [
        (r, str(ws.cell(row=r, column=H_COL).value).strip())
        for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=H_COL).value
        and str(ws.cell(row=r, column=H_COL).value).strip()
    ]
    print(f"처리 대상: {len(data_rows)}개\n")

    errors:      list[dict] = []
    issues_log:  list[dict] = []

    for i, (row_idx, original) in enumerate(data_rows, 1):
        print(f"[{i}/{len(data_rows)}] {original[:40]}")
        stage = ""
        try:
            stage = "키워드 후보 생성"
            candidates  = generate_keyword_candidates(original, keyword_model)
            stage = "카테고리 감지"
            category_id = detect_category(original, keyword_model)
            cat_name    = next((k for k, v in NAVER_CATEGORIES.items() if v == category_id), "생활/건강")
            print(f"  [1/4] 카테고리: {cat_name}")

            stage = "검색량 조회"
            search_scores   = query_search_trend(candidates, naver_id, naver_secret)
            shopping_scores = query_shopping_insight(candidates, category_id, naver_id, naver_secret)
            top_keywords    = combine_and_select(search_scores, shopping_scores, candidates)
            print(f"  [2/4] 키워드: {', '.join(top_keywords)}")

            stage = "키워드 분류"
            core_keywords, aux_words = classify_keywords(top_keywords, original, classify_model)
            print(f"  [2.5] 핵심: {core_keywords} / 보조: {aux_words}")

            stage = "상품명 최적화"
            optimized = optimize_name(original, core_keywords, aux_words, optimize_model)
            cleaned   = clean_by_rules(optimized, original)
            print(f"  [3/4] 최적화: {cleaned} ({len(cleaned)}자)")

            stage = "검수"
            final_name, issues = verify_name(original, cleaned, verify_model)
            ws.cell(row=row_idx, column=H_COL).value = final_name
            print(f"  [4/4] 최종: {final_name} ({len(final_name)}자)")

            if issues:
                issues_log.append({"행": row_idx, "원본": original, "최종": final_name, "수정사항": issues})

        except Exception as e:
            print(f"  ※ [{stage}] 오류 ({e})")
            errors.append({"행": row_idx, "원본": original, "단계": stage, "오류": str(e)})

    output_path = get_output_path(args.input)
    wb.save(output_path)
    print(f"\n저장 완료: {output_path}")
    print(f"완료: {len(data_rows) - len(errors)}건 / 오류: {len(errors)}건")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
