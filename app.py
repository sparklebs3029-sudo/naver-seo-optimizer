#!/usr/bin/env python3
"""
네이버 SEO 상품명 최적화 + 상품 소싱 - Streamlit 웹앱
"""

import base64
import hashlib
import io
import json
import os
import pathlib
import secrets
import threading
import time
from mimetypes import guess_extension
import streamlit as st
import google.generativeai as genai
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from dataclasses import dataclass, field
import extra_streamlit_components as stx

from naver_seo_agent import (
    KEYWORD_SYSTEM, CLASSIFY_SYSTEM, OPTIMIZE_SYSTEM, VERIFY_SYSTEM, GEMINI_CONFIG,
    NAVER_CATEGORIES, PROHIBITED_GROUPS,
    get_trending_products,
)
from orchestrator import run_with_orchestration, OrchestratorReport
from image_editor import image_editor as render_image_editor
from image_editor.backend import DriveUploadError, export_xlsx, load_xlsx, upload_to_drive
from image_editor.backend import delete_drive_file, fetch_image_as_b64, update_drive_file

APP_VERSION = "v1.8.0"  # 이미지 수정 탭 추가

st.set_page_config(
    page_title="셀러부스트",
    page_icon="🛒",
    layout="wide",
)

st.markdown(
    f"""
    <style>
    .block-container {{
        max-width: 96rem;
        padding-top: 1.2rem;
        padding-left: 2rem;
        padding-right: 2rem;
    }}
    iframe[title="image_editor.image_editor"] {{
        width: 100% !important;
    }}
    </style>
    <h1 style="font-size:2.5rem; line-height:1.5; padding:0.15em 0; margin-bottom:0;">셀러부스트</h1>
    <p style="color:#888; font-size:0.875rem; margin-top:0.25rem; margin-bottom:1rem;">네이버 SEO 상품명 최적화 + 트렌드 상품 소싱&nbsp;&nbsp;|&nbsp;&nbsp;{APP_VERSION}</p>
    """,
    unsafe_allow_html=True,
)

# ── 쿠키 매니저 (브라우저에 API 키 영구 저장) ────────────────────────
# @st.cache_resource 사용 금지: CookieManager는 위젯 컴포넌트이므로 매 렌더 실행 필요
_cm = stx.CookieManager(key="sb_cm")

def _save_keys(gemini: str, naver_id: str, secret: str, openai: str) -> None:
    try:
        exp = datetime.now() + timedelta(days=365)
        _cm.set("sb_gemini",       gemini,   expires_at=exp, key="save_gem")
        _cm.set("sb_naver_id",     naver_id, expires_at=exp, key="save_nid")
        _cm.set("sb_naver_secret", secret,   expires_at=exp, key="save_ns")
        _cm.set("sb_openai",       openai,   expires_at=exp, key="save_oa")
    except Exception:
        pass

def _delete_keys() -> None:
    try:
        _cm.delete("sb_gemini",       key="del_gem")
        _cm.delete("sb_naver_id",     key="del_nid")
        _cm.delete("sb_naver_secret", key="del_ns")
        _cm.delete("sb_openai",       key="del_oa")
    except Exception:
        pass

# ── 세션 상태 초기화 ─────────────────────────────────────────────────
# session_state 기본값 설정
for _k, _v in [("gemini_key", ""), ("naver_id", ""), ("naver_secret", ""), ("openai_key", ""), ("keys_saved", False)]:
    if _k not in st.session_state:
        st.session_state[_k] = _v

# 키가 아직 비어있을 때만 Secrets / 쿠키에서 로드 시도
# (extra-streamlit-components는 컴포넌트 초기화 후 자동 rerun → 그 시점에 쿠키값 수신)
if not st.session_state.gemini_key:
    # 1) Streamlit Secrets 우선
    try:
        _g = st.secrets.get("GEMINI_API_KEY", "")
        _n = st.secrets.get("NAVER_CLIENT_ID", "")
        _s = st.secrets.get("NAVER_CLIENT_SECRET", "")
        _o = st.secrets.get("OPENAI_API_KEY", "")
        if _g:
            st.session_state.gemini_key   = _g
            st.session_state.naver_id     = _n
            st.session_state.naver_secret = _s
            st.session_state.openai_key   = _o
            st.session_state.keys_saved   = True
    except Exception:
        pass

    # 2) 브라우저 쿠키 (Secrets 없을 때만)
    if not st.session_state.gemini_key:
        try:
            _all = _cm.get_all() or {}
            _g   = _all.get("sb_gemini", "") or ""
            if _g:
                st.session_state.gemini_key   = _g
                st.session_state.naver_id     = _all.get("sb_naver_id",     "") or ""
                st.session_state.naver_secret = _all.get("sb_naver_secret", "") or ""
                st.session_state.openai_key   = _all.get("sb_openai",       "") or ""
                st.session_state.keys_saved   = True
        except Exception:
            pass

# Secrets에 OPENAI_API_KEY가 없을 때 쿠키에서 보완
if not st.session_state.openai_key:
    try:
        _all = _cm.get_all() or {}
        _oa  = _all.get("sb_openai", "") or ""
        if _oa:
            st.session_state.openai_key = _oa
    except Exception:
        pass

if "daily_file_count" not in st.session_state:
    st.session_state.daily_file_count = {}

if "file_queue" not in st.session_state:
    st.session_state.file_queue = []  # [{"name": str, "bytes": bytes}, ...]

for _k, _v in [
    ("image_editor_saved_data", {}),
    ("image_editor_last_file_name", ""),
    ("image_editor_file_fingerprint", ""),
    ("image_editor_action_result", None),
    ("image_editor_ui_state", {}),
    ("active_tab", "optimizer"),
    ("image_editor_last_handled_request_id", ""),
]:
    if _k not in st.session_state:
        st.session_state[_k] = _v


def _guess_upload_ext(uploaded) -> str:
    suffix = pathlib.Path(uploaded.name).suffix.lower()
    if suffix:
        return suffix

    guessed = guess_extension(uploaded.type or "")
    return guessed or ".png"


def _build_drive_filename(prod_no: str, slot: str, uploaded) -> str:
    return f"{prod_no}_{slot}{_guess_upload_ext(uploaded)}"


def _image_summary(product: dict) -> str:
    detail_count = len(product.get("detail_imgs") or [])
    cl_state = "O" if product.get("img_cl") else "-"
    cm_state = "O" if product.get("img_cm") else "-"
    return f"{product['prod_no']} | {product['prod_name']} | CL:{cl_state} CM:{cm_state} 상세:{detail_count}"


def _build_image_action_result(action: str, ok: bool = True, **extra) -> dict:
    return {
        "action": action,
        "ok": ok,
        "request_id": extra.pop("request_id", secrets.token_hex(8)),
        **extra,
    }


def _safe_fetch_image_data(url: str) -> str | None:
    if not url:
        return None
    try:
        return fetch_image_as_b64(url)
    except Exception:
        return None


def _reset_image_editor_state() -> None:
    st.session_state.image_editor_saved_data = {}
    st.session_state.image_editor_action_result = None
    st.session_state.image_editor_ui_state = {}
    st.session_state.image_editor_last_handled_request_id = ""


def _image_file_fingerprint(file_name: str, file_bytes: bytes) -> str:
    digest = hashlib.sha1(file_bytes).hexdigest()[:16]
    return f"{file_name}:{digest}"


def _save_editor_image(prod_no: str, filename: str, data_url: str, saved_data: dict) -> dict:
    prev = saved_data.get(prod_no, {})

    if prev.get("main_file_id"):
        uploaded = update_drive_file(prev["main_file_id"], filename, data_url)
    else:
        uploaded = upload_to_drive(filename, data_url)

    next_entry = {
        **prev,
        "main_file_id": uploaded["file_id"],
        "main_url": uploaded["public_url"],
        "cl_url": uploaded["public_url"],
        "cm_url": uploaded["public_url"],
        "cl_file_id": uploaded["file_id"],
        "cm_file_id": uploaded["file_id"],
    }
    saved_data[prod_no] = next_entry

    stale_ids = {
        prev.get("main_file_id"),
        prev.get("cl_file_id"),
        prev.get("cm_file_id"),
    }
    stale_ids.discard(uploaded["file_id"])
    for file_id in stale_ids:
        if file_id:
            try:
                delete_drive_file(file_id)
            except Exception:
                pass

    return next_entry


# ── 배치 처리 상태 클래스 ──────────────────────────────────────────
@dataclass
class _BatchState:
    running:           bool               = False
    stopped:           bool               = False
    done:              bool               = False
    progress:          int                = 0
    total:             int                = 0
    log:               list               = field(default_factory=list)
    status:            str                = ""
    all_reports:       list               = field(default_factory=list)
    hard_errors:       list               = field(default_factory=list)
    result_buf:        bytes | None       = None
    out_name:          str                = ""
    stop_event:        threading.Event    = field(default_factory=threading.Event)
    total_files:       int                = 1
    current_file_idx:  int                = 1
    current_file_name: str                = ""
    file_results:      list               = field(default_factory=list)


if "batch" not in st.session_state:
    st.session_state.batch = _BatchState()

# ── 사이드바: API 키 입력 ────────────────────────────────────────────
with st.sidebar:
    st.header("API 키 설정")

    with st.expander("Gemini API 키 발급 방법"):
        st.markdown(
            "1. [Google AI Studio](https://aistudio.google.com) 접속\n"
            "2. 구글 계정으로 로그인\n"
            "3. **Get API key** 클릭 → 무료 발급"
        )

    with st.expander("네이버 API 키 발급 방법"):
        st.markdown(
            "1. [네이버 개발자센터](https://developers.naver.com) 접속\n"
            "2. 애플리케이션 등록\n"
            "3. **데이터랩(검색어트렌드)** + **데이터랩(쇼핑인사이트)** + **검색(쇼핑)** 선택\n"
            "4. Client ID · Secret 복사"
        )

    st.divider()

    gemini_key   = st.text_input("Gemini API Key",      value=st.session_state.gemini_key,   type="password", placeholder="AIzaSy...")
    naver_id     = st.text_input("Naver Client ID",     value=st.session_state.naver_id,     type="password")
    naver_secret = st.text_input("Naver Client Secret", value=st.session_state.naver_secret, type="password")
    openai_key   = st.text_input("OpenAI API Key (선택)", value=st.session_state.openai_key, type="password", placeholder="sk-...")

    keys_ready = bool(gemini_key and naver_id and naver_secret)

    col_save, col_clear = st.columns([3, 1])
    with col_save:
        if st.button("API 키 저장", use_container_width=True, type="primary", disabled=not keys_ready):
            st.session_state.gemini_key   = gemini_key
            st.session_state.naver_id     = naver_id
            st.session_state.naver_secret = naver_secret
            st.session_state.openai_key   = openai_key
            st.session_state.keys_saved   = True
            _save_keys(gemini_key, naver_id, naver_secret, openai_key)
    with col_clear:
        if st.button("삭제", use_container_width=True, disabled=not st.session_state.keys_saved):
            st.session_state.gemini_key   = ""
            st.session_state.naver_id     = ""
            st.session_state.naver_secret = ""
            st.session_state.openai_key   = ""
            st.session_state.keys_saved   = False
            _delete_keys()

    if st.session_state.keys_saved and keys_ready:
        st.success("저장됨 — 다음 접속 시 자동 로드됩니다")
        if st.session_state.openai_key:
            st.caption("OpenAI fallback 활성화: Gemini 429 시 최적화/검수 단계만 우회합니다.")
    elif not keys_ready:
        st.warning("API 키를 모두 입력 후 저장해주세요")


# ── 상단 네비게이션 ─────────────────────────────────────────────────────
tab_options = {
    "optimizer": "상품명 최적화",
    "sourcing": "상품 소싱",
    "image_editor": "이미지 수정",
}
selected_tab_label = st.radio(
    "메뉴",
    list(tab_options.values()),
    index=list(tab_options.keys()).index(st.session_state.active_tab),
    horizontal=True,
    label_visibility="collapsed",
)
selected_tab = next(key for key, value in tab_options.items() if value == selected_tab_label)
st.session_state.active_tab = selected_tab


# ════════════════════════════════════════════════════════════════════
# TAB 1: 상품명 최적화
# ════════════════════════════════════════════════════════════════════
if selected_tab == "optimizer":
    if not keys_ready:
        st.info("왼쪽 사이드바에 API 키를 먼저 입력하고 저장해주세요.")

    input_mode = st.radio("입력 방식", ["엑셀 파일", "개별 입력"], horizontal=True)

    # ── 개별 입력 모드 ────────────────────────────────────────────────
    if input_mode == "개별 입력":
        single_name = st.text_input("상품명 입력", placeholder="최적화할 상품명을 입력하세요")
        single_btn  = st.button("최적화", type="primary",
                                disabled=not keys_ready or not single_name,
                                use_container_width=True)

        if single_btn and single_name:
            genai.configure(api_key=gemini_key)
            models = {
                'keyword':  genai.GenerativeModel("gemini-2.0-flash", system_instruction=KEYWORD_SYSTEM,  generation_config=GEMINI_CONFIG),
                'classify': genai.GenerativeModel("gemini-2.0-flash", system_instruction=CLASSIFY_SYSTEM, generation_config=GEMINI_CONFIG),
                'optimize': genai.GenerativeModel("gemini-2.0-flash", system_instruction=OPTIMIZE_SYSTEM, generation_config=GEMINI_CONFIG),
                'verify':   genai.GenerativeModel("gemini-2.0-flash", system_instruction=VERIFY_SYSTEM,   generation_config=GEMINI_CONFIG),
            }
            api_keys = {'naver_id': naver_id, 'naver_secret': naver_secret, 'openai_key': openai_key}

            status_ph = st.empty()

            def single_progress(attempt: int, stage: str, detail: str = "") -> None:
                prefix = f"[시도 {attempt}/3] " if attempt > 1 else ""
                msg = f"{prefix}{stage}"
                if detail:
                    msg += f" — {detail}"
                status_ph.info(msg)

            with st.spinner("최적화 중..."):
                final_name, report = run_with_orchestration(
                    single_name, models, api_keys,
                    max_retries=3,
                    progress_callback=single_progress,
                )

            status_ph.empty()
            st.success(f"최적화 완료! (시도 {report.attempts}회)")
            st.divider()

            col_a, col_b = st.columns(2)
            with col_a:
                st.markdown("**원본 상품명**")
                st.code(single_name, language=None)
            with col_b:
                st.markdown("**최적화 상품명**")
                st.code(final_name, language=None)

            st.caption(f"글자수: {len(single_name)}자 → {len(final_name)}자")

            if report.attempts > 1:
                st.info(f"품질 기준 통과까지 {report.attempts}회 시도했습니다.")

            if report.fallback_stages:
                st.info(f"OpenAI fallback 사용 단계: {', '.join(report.fallback_stages)}")
                if report.fallback_details:
                    st.caption(" / ".join(report.fallback_details))

            if not report.passed_validation and report.warning:
                st.warning(f"오케스트레이터 경고: {report.warning}")

            if report.errors:
                with st.expander(f"처리 중 오류 {len(report.errors)}건"):
                    for err in report.errors:
                        color = "🔴" if not err.resolved else "🟠"
                        st.markdown(
                            f"{color} **[{err.stage}]** {err.error_type}  \n"
                            f"조치: {err.action_taken}  \n"
                            f"메시지: `{err.message[:100]}`"
                        )

    # ── 엑셀 파일 모드 ────────────────────────────────────────────────
    else:
        fq    = st.session_state.file_queue
        batch = st.session_state.batch

        # ── 처리 중이 아닐 때만 업로드/대기열 UI 표시 ───────────────
        if not batch.running and not ((batch.done or batch.stopped) and batch.file_results):
            # ── 파일 업로드 + 대기열 추가 ────────────────────────────
            uploaded_files = st.file_uploader(
                "엑셀 파일 업로드 (.xlsx) — 여러 파일 동시 드래그 가능",
                type=["xlsx"],
                accept_multiple_files=True,
                help="여러 파일을 한꺼번에 드래그하거나 하나씩 추가할 수 있습니다.",
                key="optimizer_file",
            )

            if uploaded_files:
                new_files    = [f for f in uploaded_files if not any(q["name"] == f.name for q in fq)]
                dup_names    = [f.name for f in uploaded_files if any(q["name"] == f.name for q in fq)]

                btn_label = f"대기열에 추가 ({len(new_files)}개)" if len(new_files) > 1 else "대기열에 추가"
                col_add, col_msg = st.columns([2, 3])
                with col_add:
                    add_btn = st.button(btn_label, type="primary", use_container_width=True,
                                        disabled=not new_files)
                with col_msg:
                    if dup_names:
                        st.warning(f"이미 추가됨: {', '.join(dup_names)}")

                if add_btn and new_files:
                    for f in new_files:
                        fq.append({"name": f.name, "bytes": f.read()})
                    st.rerun()

            # ── 대기열 표시 ──────────────────────────────────────────
            if fq:
                st.markdown(f"**대기열 ({len(fq)}개 파일)** — 순서대로 처리됩니다")
                for idx, item in enumerate(fq):
                    c_name, c_del = st.columns([5, 1])
                    c_name.text(f"{idx + 1}. {item['name']}")
                    if c_del.button("✕", key=f"del_q_{idx}", help="제거"):
                        fq.pop(idx)
                        st.rerun()

                if st.button("대기열 전체 비우기", use_container_width=True):
                    st.session_state.file_queue = []
                    st.rerun()

                st.divider()

        if fq and not batch.running and not ((batch.done or batch.stopped) and batch.file_results):
            # 열 선택은 대기열 첫 번째 파일 기준
            first_bytes = fq[0]["bytes"]
            wb_preview  = openpyxl.load_workbook(io.BytesIO(first_bytes))
            ws_preview  = wb_preview.active

            col_options = []
            for col in range(1, ws_preview.max_column + 1):
                col_letter = get_column_letter(col)
                header_val = ws_preview.cell(row=1, column=col).value
                label      = f"{col_letter}열 - {header_val}" if header_val else f"{col_letter}열"
                col_options.append(label)

            default_idx      = min(7, len(col_options) - 1)
            selected_label   = st.selectbox(
                "상품명 열 선택 (모든 파일에 적용)",
                col_options, index=default_idx,
                help="샵플링: H열 / 플레이오토: 해당 열 직접 선택"
            )
            selected_col_idx = col_options.index(selected_label) + 1

            data_rows_preview = [
                (r, str(ws_preview.cell(row=r, column=selected_col_idx).value).strip())
                for r in range(2, ws_preview.max_row + 1)
                if ws_preview.cell(row=r, column=selected_col_idx).value
                and str(ws_preview.cell(row=r, column=selected_col_idx).value).strip()
            ]
            caption = f"첫 번째 파일 기준 **{len(data_rows_preview)}개** 상품명 감지됨"
            if len(fq) > 1:
                caption += f" (파일 {len(fq)}개 전체 처리)"
            st.info(caption)

            with st.expander("상품명 미리보기 (첫 번째 파일, 상위 5개)"):
                for _, name in data_rows_preview[:5]:
                    st.text(name)

            all_raw_bytes = [item["bytes"] for item in fq]

            # ── 시작 버튼 ────────────────────────────────────────────
            start_btn = st.button("최적화 시작", type="primary",
                                  disabled=not keys_ready, use_container_width=True)
            if start_btn:
                genai.configure(api_key=gemini_key)
                models = {
                    'keyword':  genai.GenerativeModel("gemini-2.0-flash", system_instruction=KEYWORD_SYSTEM,  generation_config=GEMINI_CONFIG),
                    'classify': genai.GenerativeModel("gemini-2.0-flash", system_instruction=CLASSIFY_SYSTEM, generation_config=GEMINI_CONFIG),
                    'optimize': genai.GenerativeModel("gemini-2.0-flash", system_instruction=OPTIMIZE_SYSTEM, generation_config=GEMINI_CONFIG),
                    'verify':   genai.GenerativeModel("gemini-2.0-flash", system_instruction=VERIFY_SYSTEM,   generation_config=GEMINI_CONFIG),
                }
                api_keys = {'naver_id': naver_id, 'naver_secret': naver_secret, 'openai_key': openai_key}

                # 파일명은 메인 스레드에서 미리 계산 (스레드에서 session_state 쓰기 금지)
                now          = datetime.now()
                datetime_str = now.strftime("%Y%m%d%H%M")
                today_str    = now.strftime("%Y%m%d")
                if today_str not in st.session_state.daily_file_count:
                    st.session_state.daily_file_count = {today_str: 0}

                file_infos = []
                out_names  = []
                for item, raw in zip(fq, all_raw_bytes):
                    st.session_state.daily_file_count[today_str] += 1
                    n        = st.session_state.daily_file_count[today_str]
                    stem     = os.path.splitext(item["name"])[0]
                    out_name = f"{stem}_최적화_{datetime_str}_{n}.xlsx"
                    file_infos.append((raw, item["name"]))
                    out_names.append(out_name)

                new_batch = _BatchState(
                    running=True,
                    total_files=len(file_infos),
                    current_file_name=fq[0]["name"],
                )
                st.session_state.batch = new_batch

                def _run_batch(state: _BatchState, _file_infos, _models, _api_keys, _col_idx, _out_names):
                    total_files = len(_file_infos)
                    for file_idx, (raw_bytes, orig_name) in enumerate(_file_infos, 1):
                        if state.stop_event.is_set():
                            state.status  = f"중단됨 — 파일 {file_idx - 1}/{total_files} 완료"
                            state.stopped = True
                            break

                        state.current_file_idx  = file_idx
                        state.current_file_name = orig_name

                        _wb = openpyxl.load_workbook(io.BytesIO(raw_bytes))
                        _ws = _wb.active
                        drows = [
                            (r, str(_ws.cell(row=r, column=_col_idx).value).strip())
                            for r in range(2, _ws.max_row + 1)
                            if _ws.cell(row=r, column=_col_idx).value
                            and str(_ws.cell(row=r, column=_col_idx).value).strip()
                        ]
                        state.total    = len(drows)
                        state.progress = 0

                        for i, (row_idx, original) in enumerate(drows, 1):
                            if state.stop_event.is_set():
                                state.status  = f"중단됨 — 파일 {file_idx}/{total_files}, 행 {i - 1}/{len(drows)} 완료"
                                state.stopped = True
                                break

                            def _prog(attempt, stage, detail="", _i=i, _t=len(drows), _fi=file_idx, _ft=total_files):
                                prefix = f"[시도{attempt}] " if attempt > 1 else ""
                                state.status = f"[파일 {_fi}/{_ft}] [{_i}/{_t}] {prefix}{stage}" + (f" — {detail}" if detail else "")

                            final_name, report = run_with_orchestration(
                                original, _models, _api_keys,
                                max_retries=3,
                                progress_callback=_prog,
                            )
                            state.all_reports.append(report)
                            _ws.cell(row=row_idx, column=_col_idx).value = final_name

                            unresolved = [e for e in report.errors if not e.resolved]
                            if unresolved:
                                state.hard_errors.append({"파일": orig_name, "행": row_idx, "원본": original, "보고서": report})

                            retry_note = f" (재시도 {report.attempts}회)" if report.attempts > 1 else ""
                            warn_note  = " ⚠️" if not report.passed_validation else ""
                            fallback_note = f" [OpenAI fallback: {', '.join(report.fallback_stages)}]" if report.fallback_stages else ""
                            fallback_detail_note = ""
                            if report.fallback_details:
                                fallback_detail_note = "\n  fallback: " + " / ".join(report.fallback_details)
                            fail_note  = ""
                            if not report.passed_validation:
                                if report.validation_failures:
                                    last_fail = report.validation_failures[-1]
                                    fail_note = f"\n  사유 : {last_fail}"
                                elif report.errors:
                                    last_err = report.errors[-1]
                                    fail_note = f"\n  오류 : [{last_err.stage}] {last_err.error_type} — {last_err.message[:80]}"
                            state.log.append(
                                f"[파일{file_idx}/{total_files}][{i}/{len(drows)}]{retry_note}{warn_note}{fallback_note}\n"
                                f"  원본 : {original}\n"
                                f"  최종 : {final_name}  ({len(final_name)}자)"
                                f"{fallback_detail_note}"
                                f"{fail_note}"
                            )
                            state.progress = i

                        # 파일 1개 완료 → 결과 저장
                        buf = io.BytesIO()
                        _wb.save(buf)
                        buf.seek(0)
                        state.file_results.append({"name": _out_names[file_idx - 1], "buf": buf.read()})

                        if state.stopped:
                            break

                    if not state.stopped:
                        state.done = True
                    state.running = False

                t = threading.Thread(
                    target=_run_batch,
                    args=(new_batch, file_infos, models, api_keys, selected_col_idx, out_names),
                    daemon=True,
                )
                t.start()
                st.rerun()

        # ── 처리 중 UI ────────────────────────────────────────────
        if batch.running:
            if batch.total_files > 1:
                file_pct = (batch.current_file_idx - 1) / batch.total_files
                st.progress(file_pct, text=f"파일 {batch.current_file_idx}/{batch.total_files}: {batch.current_file_name}")
            row_pct = batch.progress / batch.total if batch.total else 0
            st.progress(row_pct, text=batch.status or "처리 준비 중...")

            if st.button("⛔ 중단", type="secondary", use_container_width=True):
                batch.stop_event.set()

            if batch.log:
                st.text_area("처리 로그", "\n\n".join(batch.log[-8:]),
                             height=300, label_visibility="collapsed")
            time.sleep(1)
            st.rerun()

        # ── 완료 / 중단 후 결과 표시 ─────────────────────────────
        if (batch.done or batch.stopped) and batch.file_results:
            all_reports   = batch.all_reports
            hard_errors   = batch.hard_errors
            success_count = sum(1 for r in all_reports if not [e for e in r.errors if not e.resolved])
            retry_count   = sum(1 for r in all_reports if r.attempts > 1)
            fallback_count = sum(1 for r in all_reports if r.fallback_stages)
            total_items   = len(all_reports)

            if batch.stopped:
                st.warning(batch.status)
            else:
                files_done = len(batch.file_results)
                st.success(
                    f"처리 완료: 파일 {files_done}개 / 상품명 {success_count}개 성공 / {len(hard_errors)}개 오류"
                    + (f" / {retry_count}개 재시도 발생" if retry_count else "")
                    + (f" / {fallback_count}개 OpenAI fallback 사용" if fallback_count else "")
                )

            st.divider()
            for fr in batch.file_results:
                st.download_button(
                    f"다운로드: {fr['name']}", data=fr["buf"],
                    file_name=fr["name"],
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary", use_container_width=True,
                )

            if st.button("새 파일 처리", use_container_width=True):
                st.session_state.batch      = _BatchState()
                st.session_state.file_queue = []
                st.rerun()

            c1, c2, c3, c4, c5 = st.columns(5)
            c1.metric("전체",   total_items)
            c2.metric("성공",   success_count)
            c3.metric("오류",   len(hard_errors))
            c4.metric("재시도", retry_count)
            c5.metric("Fallback", fallback_count)

            if hard_errors:
                st.subheader("오류 목록")
                for item in hard_errors:
                    report   = item["보고서"]
                    file_tag = f" [{item['파일']}]" if item.get("파일") else ""
                    with st.expander(f"🔴 행 {item['행']}{file_tag} | {item['원본'][:50]}"):
                        for err in report.errors:
                            icon = "🟠" if err.resolved else "🔴"
                            st.markdown(
                                f"{icon} **[{err.stage}]** {err.error_type}  \n"
                                f"조치: {err.action_taken}  \n"
                                f"메시지: `{err.message[:120]}`"
                            )
                        if report.warning:
                            st.warning(report.warning)
                        if report.fallback_stages:
                            st.info("OpenAI fallback: " + ", ".join(report.fallback_stages))
                        if report.fallback_details:
                            st.caption(" / ".join(report.fallback_details))

            retried = [r for r in all_reports if r.attempts > 1 and r.passed_validation]
            if retried:
                with st.expander(f"🟠 재시도 후 성공 {len(retried)}건"):
                    for r in retried:
                        st.markdown(
                            f"- **{r.original[:45]}** → `{r.final_name[:45]}`  "
                            f"({r.attempts}회 시도)"
                        )
                        if r.validation_failures:
                            st.caption("실패 이유: " + " / ".join(r.validation_failures[:2]))

            fallback_reports = [r for r in all_reports if r.fallback_stages]
            if fallback_reports:
                with st.expander(f"🔁 OpenAI fallback 사용 {len(fallback_reports)}건"):
                    for r in fallback_reports:
                        st.markdown(
                            f"- **{r.original[:45]}** → `{r.final_name[:45]}`  "
                            f"(단계: {', '.join(r.fallback_stages)})"
                        )
                        if r.fallback_details:
                            st.caption(" / ".join(r.fallback_details))

            if not hard_errors and retry_count == 0 and batch.done:
                st.success("특이 사항 없음. 모든 상품명이 정상 최적화되었습니다.")


# ════════════════════════════════════════════════════════════════════
# TAB 2: 상품 소싱
# ════════════════════════════════════════════════════════════════════
if selected_tab == "sourcing":
    st.subheader("네이버 쇼핑 트렌드 상품 소싱")
    st.caption("DataLab 쇼핑인사이트 + 쇼핑 검색으로 현재 잘 팔리는 상품을 찾아드립니다.")

    if not keys_ready:
        st.info("왼쪽 사이드바에 API 키를 먼저 입력하고 저장해주세요.")

    col1, col2 = st.columns(2)
    with col1:
        period_label = st.selectbox("조회 기간", ["1일 (일간)", "7일 (주간)", "30일 (월간)"], index=1)
    with col2:
        category_name = st.selectbox("카테고리", list(NAVER_CATEGORIES.keys()), index=8)

    # 금지 품목 설정
    with st.expander("금지 품목 설정", expanded=False):
        st.caption("체크된 항목에 해당하는 상품은 결과에서 자동 제외됩니다.")
        active_prohibited = []
        cols = st.columns(2)
        for idx, group in enumerate(PROHIBITED_GROUPS.keys()):
            with cols[idx % 2]:
                if st.checkbox(group, value=True, key=f"prohibit_{group}"):
                    active_prohibited.append(group)
        extra_input = st.text_input("추가 금지 키워드 (쉼표로 구분)", placeholder="예: 담배, 주류")
        extra_prohibited = [kw.strip() for kw in extra_input.split(",") if kw.strip()]

    st.divider()

    source_btn = st.button("트렌드 조회 시작", type="primary",
                           disabled=not keys_ready, use_container_width=True)

    if source_btn:
        period_days = 1 if "1일" in period_label else (7 if "7일" in period_label else 30)
        category_id = NAVER_CATEGORIES[category_name]

        with st.spinner(f"{category_name} 카테고리 트렌드를 분석 중입니다..."):
            fetched = get_trending_products(
                category_name     = category_name,
                category_id       = category_id,
                period_days       = period_days,
                client_id         = naver_id,
                client_secret     = naver_secret,
                active_prohibited = active_prohibited,
                extra_prohibited  = extra_prohibited,
                top_n             = 10,
            )

        if fetched:
            st.session_state.trend_results = fetched
        else:
            st.session_state.trend_results = []
            st.warning(
                "결과가 없습니다.\n\n"
                "**확인 사항:**\n"
                "- 네이버 앱에서 **검색(쇼핑)** API가 활성화되어 있는지 확인해주세요.\n"
                "- 네이버 개발자센터 → 내 애플리케이션 → 사용 API에 **검색** 추가 필요"
            )

    def _parse_price(price_str):
        try:
            return int(price_str.replace(",", "").replace("원", ""))
        except (ValueError, AttributeError):
            return -1

    if st.session_state.get("trend_results"):
        results = st.session_state.trend_results
        st.success(f"총 **{len(results)}개** 상품을 찾았습니다.")

        col_caption, col_pv_label, col_dropdown, col_empty = st.columns([2.5, 0.7, 1.8, 1])
        with col_caption:
            st.markdown("<p style='padding-top:7px; margin:0; color:#888; font-size:14px'>정렬 기준을 선택하세요.</p>", unsafe_allow_html=True)
        with col_pv_label:
            st.markdown("<p style='padding-top:7px; margin:0'>판매가</p>", unsafe_allow_html=True)
        with col_dropdown:
            price_sort = st.selectbox(
                "판매가",
                ["높은 가격순", "낮은 가격순"],
                key="price_sort",
                label_visibility="collapsed",
            )

        # 키워드별로 그룹화 (트렌드 점수 내림차순 고정)
        from collections import defaultdict
        keyword_order = []
        keyword_groups = defaultdict(list)
        for item in sorted(results, key=lambda x: x["트렌드점수"], reverse=True):
            if item["키워드"] not in keyword_groups:
                keyword_order.append((item["키워드"], item["트렌드점수"]))
            keyword_groups[item["키워드"]].append(item)

        # 각 그룹 내 판매가 정렬
        results_sorted = []
        for keyword, score in keyword_order:
            group = keyword_groups[keyword]
            if price_sort == "높은 가격순":
                group = sorted(group, key=lambda x: _parse_price(x["최저가"]), reverse=True)
            elif price_sort == "낮은 가격순":
                group = sorted(group, key=lambda x: (_parse_price(x["최저가"]) == -1, _parse_price(x["최저가"])))
            st.markdown(f"#### 키워드: {keyword}  (트렌드 점수: {score})")
            for item in group:
                col_a, col_b, col_c = st.columns([4, 2, 2])
                with col_a:
                    st.markdown(f"[{item['상품명']}]({item['링크']})")
                with col_b:
                    st.write(item["최저가"])
                with col_c:
                    st.write(item["쇼핑몰"])
            results_sorted.extend(group)

        st.divider()

        with st.expander("전체 결과 표로 보기"):
            import pandas as pd
            df = pd.DataFrame(results_sorted)[["키워드", "트렌드점수", "상품명", "최저가", "쇼핑몰", "카테고리"]]
            st.dataframe(df, use_container_width=True, hide_index=True)


# ════════════════════════════════════════════════════════════════════
# TAB 3: 이미지 수정
# ════════════════════════════════════════════════════════════════════
if selected_tab == "image_editor":
    st.markdown(
        """
        <style>
        h1, div[data-testid="stCaptionContainer"] {
            display: none !important;
        }
        .block-container {
            padding-top: 0.6rem !important;
        }
        div[role="radiogroup"] {
            margin-top: 0 !important;
            margin-bottom: 0.35rem !important;
        }
        div[data-testid="stFileUploader"] > label,
        div[data-testid="stFileUploaderDropzoneInstructions"],
        div[data-testid="stFileUploaderDropzone"] small {
            font-size: 0.92rem !important;
        }
        div[data-testid="stFileUploaderDropzone"] {
            padding-top: 0.55rem !important;
            padding-bottom: 0.55rem !important;
        }
        div[data-testid="stFileUploader"] {
            margin-bottom: 0.15rem !important;
        }
        div[data-testid="stHorizontalBlock"] p {
            margin-bottom: 0 !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    uploader_col, status_col = st.columns([1.35, 1])
    with uploader_col:
        uploaded_image_file = st.file_uploader(
            "이미지 수정용 엑셀 업로드 (.xlsx)",
            type=["xlsx"],
            key="image_editor_file",
            help="A열 상품번호 기준으로 90열(CL), 91열(CM) 이미지 URL을 갱신합니다.",
        )
    with status_col:
        saved_count = len(st.session_state.image_editor_saved_data)
        st.caption(f"현재 반영 대기 상품 수: {saved_count}개" if saved_count else " ")
        if st.button("이미지 수정 작업 초기화", use_container_width=True, disabled=not saved_count):
            st.session_state.image_editor_saved_data = {}
            st.session_state.image_editor_action_result = None
            st.session_state.image_editor_ui_state = {}
            st.rerun()

    if uploaded_image_file:
        image_xlsx_bytes = uploaded_image_file.read()
        st.session_state.image_editor_last_file_name = uploaded_image_file.name
        current_fingerprint = _image_file_fingerprint(uploaded_image_file.name, image_xlsx_bytes)
        if st.session_state.image_editor_file_fingerprint != current_fingerprint:
            _reset_image_editor_state()
            st.session_state.image_editor_file_fingerprint = current_fingerprint

        try:
            products, _row_map = load_xlsx(image_xlsx_bytes)
        except Exception as exc:
            st.error(f"엑셀을 읽는 중 오류가 발생했습니다: {exc}")
            products = []

        if products:
            editable_products = [
                product for product in products
                if product.get("prod_no") and (
                    product.get("img_cl") or product.get("img_cm") or product.get("detail_imgs")
                )
            ]

            st.caption(
                f"총 {len(products)}개 상품을 읽었습니다. "
                f"이미지 정보가 있는 상품은 {len(editable_products)}개입니다."
            )

            if editable_products:
                component_action = render_image_editor(
                    products=editable_products,
                    action_result=st.session_state.image_editor_action_result,
                    saved_data=st.session_state.image_editor_saved_data,
                    ui_state=st.session_state.image_editor_ui_state,
                    key="image_editor_component",
                )

                if component_action:
                    request_id = component_action.get("request_id", secrets.token_hex(8))
                    if request_id != st.session_state.image_editor_last_handled_request_id:
                        st.session_state.image_editor_last_handled_request_id = request_id
                        st.session_state.image_editor_ui_state = component_action.get("ui_state", {})
                        action = component_action.get("action")

                        try:
                            if action == "fetch_image":
                                prod_no = str(component_action.get("prod_no", ""))
                                product = next((item for item in editable_products if str(item.get("prod_no")) == prod_no), None)
                                data_url = fetch_image_as_b64(component_action["url"])
                                st.session_state.image_editor_action_result = _build_image_action_result(
                                    "fetch_image",
                                    request_id=request_id,
                                    prod_no=prod_no,
                                    url=component_action.get("url"),
                                    thumb_idx=component_action.get("thumb_idx", 0),
                                    data_url=data_url,
                                    compare_cl_data_url=_safe_fetch_image_data((product or {}).get("img_cl", "")),
                                    compare_cm_data_url=_safe_fetch_image_data((product or {}).get("img_cm", "")),
                                )
                            elif action == "save_image":
                                prod_no = str(component_action["prod_no"])
                                filename = component_action.get("filename") or f"{prod_no}_main.jpg"
                                saved_entry = _save_editor_image(
                                    prod_no,
                                    filename,
                                    component_action["data_url"],
                                    st.session_state.image_editor_saved_data,
                                )
                                current_idx = int(st.session_state.image_editor_ui_state.get("selected_index", 0))
                                next_idx = min(current_idx + 1, len(editable_products) - 1)
                                st.session_state.image_editor_action_result = _build_image_action_result(
                                    "save_image",
                                    request_id=request_id,
                                    filename=filename,
                                    prod_no=prod_no,
                                    preview_data_url=component_action.get("preview_data_url"),
                                    cl_url=saved_entry.get("cl_url"),
                                    cm_url=saved_entry.get("cm_url"),
                                    next_index=next_idx if current_idx < len(editable_products) - 1 else current_idx,
                                    saved_data=st.session_state.image_editor_saved_data,
                                )
                            elif action == "export_xlsx":
                                out_name = f"{pathlib.Path(uploaded_image_file.name).stem}_이미지수정.xlsx"
                                file_bytes = export_xlsx(image_xlsx_bytes, st.session_state.image_editor_saved_data)
                                export_result = _build_image_action_result(
                                    "export_xlsx",
                                    request_id=request_id,
                                    filename=out_name,
                                    file_b64=base64.b64encode(file_bytes).decode("ascii"),
                                )
                                _reset_image_editor_state()
                                st.session_state.image_editor_action_result = export_result
                            st.rerun()
                        except DriveUploadError as exc:
                            st.session_state.image_editor_action_result = _build_image_action_result(
                                action or "unknown",
                                ok=False,
                                request_id=request_id,
                                error=str(exc),
                            )
                            st.rerun()
                        except Exception as exc:
                            st.session_state.image_editor_action_result = _build_image_action_result(
                                action or "unknown",
                                ok=False,
                                request_id=request_id,
                                error=str(exc),
                            )
                            st.rerun()

            else:
                st.warning("엑셀에서 이미지 정보가 있는 상품을 찾지 못했습니다.")
        else:
            st.warning("표시할 상품 데이터가 없습니다.")
    else:
        st.markdown(
            "**사용 방법**\n"
            "1. 샵플링 엑셀 파일 업로드\n"
            "2. 상품 선택 후 CL 또는 CM 이미지 업로드\n"
            "3. Drive 업로드 완료 후 결과 엑셀 다운로드"
        )
